import numpy as np
import pandas as pd
import math
import matplotlib.pyplot as plt
from tqdm import tqdm
import openpyxl
FileName = "DF_PercentageError.csv"

# Layer making class
class SetLayer:
    def __init__(self, Length, ThermalConductivity, VolumetricHeatCapacity, NodeNum):
        self.Length = Length # 레이어 길이 
        self.Cell_length = Length/NodeNum # 셀 길이
        self.ThermalConductivity = ThermalConductivity # 열전도율
        self.VolumetricHeatCapacity = VolumetricHeatCapacity # 체적열용량
        self.NodeNum = NodeNum # 차분 수
        self.ThermalResistance = (Length/NodeNum)/ThermalConductivity # 열저항
        self.ThermalConductance = ThermalConductivity/(Length/NodeNum) # 
        self.ThermalDiffusivity = ThermalConductivity/(VolumetricHeatCapacity)

    def L(self): 
        return self.Length
    def dx(self): 
        return self.Cell_length
    def k(self): 
        return self.ThermalConductivity
    def C(self): 
        return self.VolumetricHeatCapacity
    def n(self): 
        return self.NodeNum
    def R(self):
        return self.ThermalResistance
    def K(self):
        return self.ThermalConductance
    def TD(self):
        return self.ThermalDiffusivity
    
# Construction = [[Layer 1],[Layer 2],[Layer 3]]

def NumOfCellsConstruction(Construction): 
    return sum([Construction[i].n() for i in range(len(Construction))])

def NumOfCellsLayer(Layer): 
    return Layer.n()

def ResistanceOfConstruction(Construction):
    return sum([Construction[i].R()*NumOfCellsLayer(Construction[i]) for i in range(len(Construction))])

def ConductanceOfConstruction(Construction):
    return 1/ResistanceOfConstruction(Construction)

def D2K(Temperature): #Degree to Kelvin
    return Temperature + 273.15
def K2D(Temperature): #Kelvin to Degree
    return Temperature - 273.15
def NoneMatrix(rows, cols): #Make None matrix
    return [[None]*cols for _ in range(rows)]  
def NoneList(N): #Make None List
    return [None]*N
def cm2in(value):
    return value/2.54

def LinearInterpolationList(x1,x2,divide):
    InterpolatedList = []
    grad = (x2-x1)/(divide-1)
    for i in range(divide-2):
        InterpolatedList.append(x1+grad*(i+1))
    InterpolatedList.insert(0,x1)
    InterpolatedList.append(x2)
    return InterpolatedList

# Unit change 
day_to_hour = 24
hour_to_min = 60
hour_to_sec = 3600
min_to_sec = 60  
m2cm = 100
cm2m = 1/100

# Time variable
TimeStep = 10 #[s] 
PSTime = 120 #PS:PreSimulation[h]
TSTime = 144+1 #TS:DFSimulation[h]
Duration = TSTime - PSTime #[h]
NumOfTimeStep = int(TSTime*hour_to_sec/TimeStep)
TimeStepList = [TimeStep*i for i in range(NumOfTimeStep+1)] #time step list [t1,t2,..., tN]
time = pd.DataFrame(TimeStepList) #second
time_min = time/60; #minute
time_hour = time/3600; #hour
time_end = time_hour[-1:]; #last time

PSTimeRowNum = int(PSTime*hour_to_sec/TimeStep)
TSTimeRowNum = int((TSTime)*hour_to_sec/TimeStep)

# Calculation condition
InitialTemp = D2K(20) 
TBC1 = [(D2K(20 + 10*(math.sin(2*math.pi*n/(day_to_hour*hour_to_sec))))) for n in TimeStepList] # [K]
TBC2 = [(D2K(20)) for _ in TimeStepList]

LeftBC = TBC1 #Left boundary condition
RightBC = TBC2 #Right boundary condition
T0 = TBC1 #Environment temperature

# Material properties
TDUnit= 10**(-6) #TD:Thermal Diffusivity
MinTD = 0.5*TDUnit #TD:Thermal Diffusivity
MaxTD = 5*TDUnit #TD:Thermal Diffusivity
TDInterval = 1.5*TDUnit #TD:Thermal Diffusivity
# TDList =  np.arange(MinTD, MaxTD+TDInterval, TDInterval)
TDList = [0.5*TDUnit, 1.0*TDUnit, 2.0*TDUnit]
NumOfTD = len(TDList)

TcList = [1 for _ in range(NumOfTD)] #TcList:ThermalConductivity 
VHC = [TcList[i]/TDList[i] for i in range(NumOfTD)] #VHC:VolumetricHeatcListapacity

# System Length
MinLength = 0.05 #[m]
MaxLength = 0.35 #[m]
LengthInterval = 0.15 #[m]
# SystemLengthList = np.arange(MinLength,MaxLength+LengthInterval,LengthInterval)
SystemLengthList = [0.05,0.1,0.2]
NumOfLength = len(SystemLengthList)

DiscretizedLength = 0.01 #[m]
NodeNumList = [round(SystemLengthList[i]/DiscretizedLength) for i in range(NumOfLength)]

# DataFrame to save
x = SystemLengthList
y = TDList
x_pos, y_pos = np.meshgrid(x, y)
DF_PercentageError = pd.DataFrame(np.zeros((NumOfTD,NumOfLength)))

# Make Excel file to save
for TD_idx in range(NumOfTD): #ThermalDiffusivity(y axis)
    for len_idx in range(NumOfLength): #Length(x axis)
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_XcRate.xlsx')
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_XcRate.xlsx')
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_CellXcRate.xlsx')
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_CellXcRate.xlsx')
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_T.xlsx')
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_T.xlsx')
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_q.xlsx')
        wb = openpyxl.Workbook()
        wb.save(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_q.xlsx')

# Run
for TD_idx in range(NumOfTD): #ThermalDiffusivity(y axis)
    print(f"Thermal diffusivity : {TDList[TD_idx]}")
    for len_idx in range(NumOfLength): #Length(x axis)
        print(f"Length : {round(NodeNumList[len_idx])} cm")
        # Define the thermal network
        Layer = SetLayer(Length=SystemLengthList[len_idx], 
                         ThermalConductivity=TcList[TD_idx], 
                         VolumetricHeatCapacity=VHC[TD_idx],
                         NodeNum=NodeNumList[len_idx])
        Construction = [Layer]
        N = NumOfCellsConstruction(Construction)
        dx = [Lidx.dx() for Lidx in Construction for _ in range(NumOfCellsLayer(Lidx))]
        dx_L, dx_R = [(dx[i-1] + dx[i])/2 for i in range(1,N)], [(dx[i] + dx[i+1])/2 for i in range(N-1)]
        dx_L.insert(0,dx[0]/2)
        dx_R.append(dx[N-1]/2)
        k = [Lidx.k() for Lidx in Construction for _ in range(NumOfCellsLayer(Lidx))]
        C = [Lidx.C() for Lidx in Construction for _ in range(NumOfCellsLayer(Lidx))]
        R = [Lidx.R() for Lidx in Construction for _ in range(NumOfCellsLayer(Lidx))]
        K = [Lidx.K() for Lidx in Construction for _ in range(NumOfCellsLayer(Lidx))]
        K_L, K_R = [K[i] for i in range(1,N)], [K[i] for i in range(N-1)]
        K_L.insert(0,K[0]*2)
        K_R.append(K[N-1]*2)

        # Empty list and matrix for data saving
        rows, cols = NumOfTimeStep, N
        T_L, T, T_R = NoneMatrix(rows,cols), NoneMatrix(rows,cols), NoneMatrix(rows,cols)
        q_in, q, q_out = NoneMatrix(rows,cols), NoneMatrix(rows,cols), NoneMatrix(rows,cols)
        T_L_half, T_half, T_R_half = NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols)
        Carneff = NoneMatrix(rows-1,cols)
        T0_half = NoneList(rows-1)
        q_in_half, q_half, q_out_half = NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols)
        U_CellXifRate, U_CellXcRate, U_CellXstRate, U_CellXofRate, U_CellstX = NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols), NoneMatrix(rows-1,cols)
        D_XifRate, D_XcRate, D_XofRate = NoneList(rows), NoneList(rows), NoneList(rows)

        # Initial Temperature, boundary condition
        for i in range(N):
            T_L[0][i], T[0][i], T_R[0][i] = InitialTemp, InitialTemp, InitialTemp
        for n in range(NumOfTimeStep):
            T_L[n][0], T_R[n][N-1] = LeftBC[n], RightBC[n]

        # TDMA(Triangle Diagonal Matrix Algorithm)
        aList = [-TimeStep*K_L[i] for i in range(N)]
        bList = [2*dx[i]*C[i]+TimeStep*K_L[i]+TimeStep*K_R[i] for i in range(N)]
        cList = [-TimeStep*K_R[i] for i in range(N)]

        # AMatrix T = BMatrix
        AMatrix = np.zeros((N, N))

        for idx in range(N-1):
            AMatrix[idx+1][idx] = aList[idx+1]
            AMatrix[idx][idx]   = bList[idx]
            AMatrix[N-1][N-1]   = bList[N-1]
            AMatrix[idx][idx+1] = cList[idx]
            
        AMatrixInversed = np.linalg.inv(AMatrix)

        for n in range(NumOfTimeStep-1):  
            BMatrix = [] 
            for i in range(1,N-1): #Node
                gi = [TimeStep*K_L[i]*T[n][i-1]+(2*dx[i]*C[i]-TimeStep*K_L[i]-TimeStep*K_R[i])*T[n][i]+TimeStep*K_R[i]*T[n][i+1]] 
                BMatrix.append(gi)
            g1 = [2*TimeStep*K_L[0]*T_L[n][0]+(2*dx[0]*C[0]-TimeStep*K_L[0]-TimeStep*K_R[0])*T[n][0]+TimeStep*K_R[0]*T[n][1]]
            gN = [TimeStep*K_L[N-1]*T[n][N-2]+(2*dx[N-1]*C[N-1]-TimeStep*K_L[N-1]-TimeStep*K_R[N-1])*T[n][N-1]+2*TimeStep*K_R[N-1]*T_R[n][N-1]]
            BMatrix.insert(0,g1)  
            BMatrix.append(gN)

            # Temperature
            for i in range(N): #Get next time step temperature value
                T[n+1][i] = np.dot(AMatrixInversed, BMatrix)[i][0]
            for i in range(1,N): #Define left interface temperature
                T_L[n+1][i] = (T[n+1][i-1]+T[n+1][i])/2
            for i in range(0,N-1): #Define right interface temperature
                T_R[n+1][i] = (T[n+1][i]+T[n+1][i+1])/2

        # Heatflux
        for n in range(NumOfTimeStep):
            for i in range(1,N): #Heat influx
                q_in[n][i] = K_L[i]*(T[n][i-1] - T[n][i])
            q_in[n][0] = K_L[0]*(T_L[n][0]-T[n][0]) #Heat influx 1st half node
            for i in range(N-1): #Heat outflux 
                q_out[n][i] = K_R[i]*(T[n][i] - T[n][i+1])
            q_out[n][N-1] = K_R[N-1]*(T[n][N-1]-T_R[n][N-1]) #Heat outflux last half node
            for i in range(N): #Interpolated Heatflux for exergy calculation
                q[n][i] = (q_in[n][i] + q_out[n][i])/2

        # Define half time step values
        for n in range(NumOfTimeStep-1):
            for i in range(N):
                #Temperature
                T_L_half[n][i] = (T_L[n][i] + T_L[n+1][i])/2
                T_half[n][i] = (T[n][i] + T[n+1][i])/2
                T_R_half[n][i] = (T_R[n][i] + T_R[n+1][i])/2
                T0_half[n] = (T0[n]+T0[n+1])/2  
                #Heat flux
                q_in_half[n][i] = (q_in[n][i] + q_in[n+1][i])/2
                q_half[n][i] = (q[n][i] + q[n+1][i])/2
                q_out_half[n][i] = (q_out[n][i] + q_out[n+1][i])/2

        # Define unsteady-state exergy values
        for n in range(NumOfTimeStep-1):
            for i in range(N):
                U_CellXifRate[n][i] = q_in_half[n][i]*(1-(T0_half[n]/T_L_half[n][i])) #Exergy inflow rate
                U_CellXcRate[n][i] = (dx[i]/k[i])*(T0_half[n]*(q_half[n][i]/T_half[n][i])**2) #Exergy consumption rate
                U_CellXstRate[n][i] = dx[i]*C[i]*(1-(T0_half[n]/T_half[n][i]))*((T[n+1][i]-T[n][i])/TimeStep) #Exergy stored rate
                U_CellXofRate[n][i] = q_out_half[n][i]*(1-(T0_half[n]/T_R_half[n][i])) #Exergy outflow rate
                U_CellstX[n][i] = C[i]*dx[i]*((T[n][i]-T0[n])-T0[n]*math.log(T[n][i]/T0[n])) #Stored exergy in cell

    # Define steady-state exergy values (Dynamic)

        #Temperature (처음과 끝노드 온도 계산후 선형보간)
        D_T = []
        for n in range(NumOfTimeStep):
            TempGradiant = (RightBC[n]-LeftBC[n])/(SystemLengthList[len_idx]) #n timestep gradiant
            FNTemp = LeftBC[n]+TempGradiant*(DiscretizedLength/2) #First node temperature
            LNTemp = RightBC[n]-TempGradiant*(DiscretizedLength/2) #Last node temperature
            D_T.append(LinearInterpolationList(FNTemp,LNTemp,NodeNumList[len_idx])) #n time step에서의
        DF_D_T = K2D(pd.DataFrame(D_T)[PSTimeRowNum:TSTimeRowNum]) #Dataframing

        #Heatflux
        D_q_tot = [ConductanceOfConstruction(Construction)*(LeftBC[n] - RightBC[n]) for n in range(NumOfTimeStep)] #Total heat flux 
        DF_D_q_tot = pd.DataFrame(D_q_tot)
        for i in range(NodeNumList[len_idx]): #Heat flux copy for imaginary discretization (as many Node)
            DF_D_q_tot[i] = DF_D_q_tot[0]
        DF_D_q = DF_D_q_tot[PSTimeRowNum:TSTimeRowNum]
        
        #Exergy
        for n in range(NumOfTimeStep):
            D_XifRate[n]  = (1-(T0[n]/LeftBC[n]))*D_q_tot[n] #Exergy inflow rate
            D_XcRate[n]  = ResistanceOfConstruction(Construction)*(D_q_tot[n]**2/(LeftBC[n]*RightBC[n]))*T0[n] #Exergy consumption rate
            D_XofRate[n] = (1-(T0[n]/RightBC[n]))*D_q_tot[n] #Exergy outflow rate
        
        D_CellXcRate = []
        for n in range(NumOfTimeStep):
            D_CellXcRate.append([D_XcRate[n]/N for _ in range(N)]) #Imaginary discretization
        DF_D_CellXcRate = pd.DataFrame(D_CellXcRate)[PSTimeRowNum:TSTimeRowNum]

        # Carnot coefficient 
        for n in range(NumOfTimeStep-1):
            for i in range(N):
                Carneff[n][i] = 1-T0[n]/T_half[n][i]

    # Data framing
        # Exergy consumption
        DF_U_CellXcRate = pd.DataFrame(U_CellXcRate)
        DF_U_XcRate = DF_U_CellXcRate.sum(axis=1) #Spatially summed up the exergy consumption rate
        DF_U_XcRate = DF_U_XcRate.iloc[PSTimeRowNum:TSTimeRowNum] #Remove pre-simulation data\

        DF_D_XcRate = pd.DataFrame(D_XcRate)
        DF_D_XcRate = DF_D_XcRate.iloc[PSTimeRowNum:TSTimeRowNum] #Remove pre-simulation data
        
        DF_U_Xc = DF_U_XcRate.sum(axis=0)*TimeStep #Time dimensionary summed up the exergy consumption rate
        DF_D_Xc = DF_D_XcRate.sum(axis=0)*TimeStep #Time dimensionary summed up the exergy consumption rate

        # Percentage error
        PercentageError = round(abs((DF_U_Xc-DF_D_Xc)/DF_U_Xc)*100,1) #Exergy consumption percentage error
        print(f"PercentageError: {PercentageError}%")
        DF_PercentageError.iloc[TD_idx,len_idx] = PercentageError #Save in dataframe

        # Temperature
        DF_T_L  = K2D(pd.DataFrame(T_L)[PSTimeRowNum:TSTimeRowNum]) #Temperature kelvin to degree
        DF_T_R  = K2D(pd.DataFrame(T_R)[PSTimeRowNum:TSTimeRowNum]) #Temperature kelvin to degree
        DF_T    = K2D(pd.DataFrame(T)[PSTimeRowNum:TSTimeRowNum]) #Temperature kelvin to degree
        
        # Heatflux
        DF_q_in  = pd.DataFrame(q_in)[PSTimeRowNum:TSTimeRowNum]
        DF_q_out = pd.DataFrame(q_out)[PSTimeRowNum:TSTimeRowNum]
        DF_q    = pd.DataFrame(q)[PSTimeRowNum:TSTimeRowNum]
        
        # Carnot coefficient
        DF_Carneff = pd.DataFrame(Carneff)[PSTimeRowNum:TSTimeRowNum]

        # Unsteady state 
        DF_U_CellXifRate = pd.DataFrame(U_CellXifRate)[PSTimeRowNum:TSTimeRowNum]
        DF_U_CellXcRate = pd.DataFrame(U_CellXcRate)[PSTimeRowNum:TSTimeRowNum]
        DF_U_CellXstRate = pd.DataFrame(U_CellXstRate)[PSTimeRowNum:TSTimeRowNum]
        DF_U_CellXofRate = pd.DataFrame(U_CellXofRate)[PSTimeRowNum:TSTimeRowNum]
        DF_U_CellstX = pd.DataFrame(U_CellstX)[PSTimeRowNum:TSTimeRowNum]

        DF_U_XifRate  = DF_U_CellXifRate.sum(axis=1).to_frame()[PSTimeRowNum:TSTimeRowNum]
        DF_U_XstRate  = DF_U_CellXstRate.sum(axis=1).to_frame()[PSTimeRowNum:TSTimeRowNum]
        DF_U_XofRate = DF_U_CellXofRate.sum(axis=1).to_frame()[PSTimeRowNum:TSTimeRowNum]

        DF_D_XifRate = pd.DataFrame(D_XifRate)[PSTimeRowNum:TSTimeRowNum]
        DF_D_XofRate = pd.DataFrame(D_XofRate)[PSTimeRowNum:TSTimeRowNum]

    # Save to Excel
        #unsteady   
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_XcRate.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_U_XcRate.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_CellXcRate.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_U_CellXcRate.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_T.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_T.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_U_q.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_q.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index
            
        #dynamic
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_CellXcRate.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_D_CellXcRate.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_XcRate.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_D_XcRate.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_T.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_D_T.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index
        with pd.ExcelWriter(f'alpha={TDList[TD_idx]*(1/TDUnit)}/DF_D_q.xlsx', engine="openpyxl", mode='a') as writer: #Distinguish file using TD index
            DF_D_q.to_excel(writer, sheet_name=f'{round(NodeNumList[len_idx])} cm', index=False) #Distinguish sheet using the length index

# Save to csv
print("END")